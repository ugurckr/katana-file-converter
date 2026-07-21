"""Excel dönüştürücüleri: csv <-> xlsx, json -> xlsx.

openpyxl ile çalışır; Excel kurulumu gerekmez.
"""

import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .base import register


def _rows_to_xlsx(header: list[str], rows: list[list], dst: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dst))


@register(".csv", ".xlsx", "Excel tablosu")
def csv_to_xlsx(src: Path, dst: Path) -> None:
    with src.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"'{src}' boş, dönüştürülecek satır yok.")
        _rows_to_xlsx(header, list(reader), dst)


@register(".xlsx", ".csv", "CSV tablo (ilk sayfa)")
def xlsx_to_csv(src: Path, dst: Path) -> None:
    wb = load_workbook(str(src), read_only=True, data_only=True)
    try:
        ws = wb.active
        dst.parent.mkdir(parents=True, exist_ok=True)
        with dst.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
    finally:
        wb.close()


@register(".json", ".xlsx", "Excel tablosu")
def json_to_xlsx(src: Path, dst: Path) -> None:
    data = json.loads(src.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list) or not all(isinstance(row, dict) for row in data):
        raise ValueError("JSON, bir obje listesi (veya tek bir obje) olmalı.")
    header = list(dict.fromkeys(key for row in data for key in row))
    rows = [
        [json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v for v in
         (row.get(key, "") for key in header)]
        for row in data
    ]
    _rows_to_xlsx(header, rows, dst)
